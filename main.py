import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment
from win32com.client import Dispatch
from sklearn.cluster import KMeans
import matplotlib.pyplot as plt
from io import BytesIO
from openpyxl.drawing.image import Image  # <-- Needed for inserting images

# Set directory paths
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.path.join(BASE_DIR, "data")
REPORT_DIR = os.path.join(BASE_DIR, "reports")

def load_and_prepare_data():
    customers = pd.read_csv(os.path.join(DATA_DIR, 'customers.csv'))
    orders = pd.read_csv(os.path.join(DATA_DIR, 'orders.csv'))
    products = pd.read_csv(os.path.join(DATA_DIR, 'products.csv'))

    products['category'] = products['category'].replace('Electronics', 'Electronics')

    merged = pd.merge(
        pd.merge(orders, customers, left_on='customer_id', right_on='id'),
        products, left_on='id_y', right_on='id'
    )

    return merged

def apply_ai_analysis(df):
    rfm = df.groupby('customer_id').agg({
        'amount': ['sum', 'count'],
        'order_date': lambda x: (pd.to_datetime('today') - pd.to_datetime(x.max())).days
    })
    rfm.columns = ['monetary', 'frequency', 'recency']

    kmeans = KMeans(n_clusters=3)
    df['segment'] = df['customer_id'].map(
        pd.Series(kmeans.fit_predict(rfm), index=rfm.index)
    )

    df['amount_zscore'] = (df['amount'] - df['amount'].mean()) / df['amount'].std()
    df['is_anomaly'] = df['amount_zscore'].abs() > 2

    return df

def create_excel_report(df):
    wb = Workbook()

    # Data sheet
    ws_data = wb.active
    ws_data.title = "Transaction Data"
    for r in dataframe_to_rows(df, index=False, header=True):
        ws_data.append(r)

    for cell in ws_data[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    # Summary sheet
    ws_summary = wb.create_sheet("Summary")
    metrics = {
        'Total Customers': df['customer_id'].nunique(),
        'Total Revenue': f"${df['amount'].sum():,.2f}",
        'Avg Order Value': f"${df['amount'].mean():.2f}",
        'Top Product': df['name_y'].mode()[0]
    }
    for i, (k, v) in enumerate(metrics.items(), start=1):
        ws_summary[f'A{i}'] = k
        ws_summary[f'B{i}'] = v
        ws_summary[f'A{i}'].font = Font(bold=True)

    # Charts
    ws_charts = wb.create_sheet("Charts")
    fig, ax = plt.subplots()
    df.groupby('segment')['amount'].sum().plot(kind='bar', ax=ax)
    ax.set_title('Revenue by Customer Segment')
    chart_buffer = BytesIO()
    plt.savefig(chart_buffer, format='png')
    plt.close()
    chart_buffer.seek(0)
    img = Image(chart_buffer)
    ws_charts.add_image(img, 'A1')

    report_path = os.path.join(REPORT_DIR, 'sales_report.xlsx')
    wb.save(report_path)
    return report_path

def open_in_excel(file_path):
    xl = Dispatch("Excel.Application")
    xl.Visible = True
    wb = xl.Workbooks.Open(file_path)
    return xl, wb

if __name__ == "__main__":
    print("Loading data...")
    data = load_and_prepare_data()

    print("Applying AI analysis...")
    enhanced_data = apply_ai_analysis(data)

    print("Generating Excel report...")
    report_file = create_excel_report(enhanced_data)

    print("Opening report...")
    excel_app, workbook = open_in_excel(os.path.abspath(report_file))

    print(f"Report generated at: {os.path.abspath(report_file)}")
